"""Turn an uploaded file of any supported format into rows of cells.

Everything downstream — header detection, column aliases, date and amount
parsing, fingerprinting, the rules, the model — works on a list of rows.
So supporting a new format means writing one function that produces that
list, and nothing else in the app changes.

Supported: CSV (and other delimited text), JSON, Excel .xlsx (every sheet).

The contract is the same as the parser's: never raise on messy content.
Only a file that cannot be read as a table at all raises, and that error
carries whatever was recognised so the UI can say what was wrong.
"""

import csv
import io
import json

ENCODINGS = ("utf-8-sig", "utf-8", "latin-1")
DELIMITERS = ",;\t|"

# JSON files often wrap the rows in a envelope. These are the usual names.
LIST_KEYS = ("transactions", "data", "rows", "items", "records", "results")

EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
JSON_EXTENSIONS = (".json",)

# The first bytes of any zip archive, which is what an .xlsx actually is.
ZIP_MAGIC = b"PK\x03\x04"


class UnreadableFile(Exception):
    """The file could not be read as a table at all."""

    def __init__(self, message, detected_columns=None):
        super().__init__(message)
        self.detected_columns = detected_columns or []


def detect_format(file_bytes: bytes, filename=None) -> str:
    """Return 'excel', 'json' or 'delimited'.

    The extension is a hint, not proof — people rename files, and browsers
    lie about content types — so the bytes get the final say where they are
    unambiguous.
    """
    if file_bytes.startswith(ZIP_MAGIC):
        return "excel"

    name = (filename or "").lower()
    if name.endswith(EXCEL_EXTENSIONS):
        return "excel"
    if name.endswith(JSON_EXTENSIONS):
        return "json"

    if _looks_like_json(file_bytes):
        return "json"

    return "delimited"


def _looks_like_json(file_bytes: bytes) -> bool:
    head = file_bytes.lstrip()[:1]
    return head in (b"{", b"[")


def decode(file_bytes: bytes) -> str:
    """Text from bytes, trying the encodings Indian bank exports use."""
    for encoding in ENCODINGS:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnreadableFile("File is not readable as text.")


# --- delimited text -------------------------------------------------------

def read_delimited(file_bytes: bytes):
    """CSV and its relatives. The delimiter is sniffed, not assumed."""
    text = decode(file_bytes)

    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=DELIMITERS)
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


# --- json -----------------------------------------------------------------

def _records_from_json(payload):
    """Find the list of transactions inside whatever shape the file has."""
    if isinstance(payload, list):
        return payload

    if isinstance(payload, dict):
        for key in LIST_KEYS:
            if isinstance(payload.get(key), list):
                return payload[key]
        # A dict of dicts, keyed by id, is also a reasonable export shape.
        values = list(payload.values())
        if values and all(isinstance(value, dict) for value in values):
            return values

    raise UnreadableFile(
        "JSON must be a list of transactions, or an object containing one "
        f"under a key such as {LIST_KEYS[0]!r}."
    )


def _columns_in_order(records):
    """Every key seen, in the order the file first mentions it.

    Not just the first record's keys: exports routinely omit a null field on
    some rows, and dropping a column because row 1 lacked it loses data.
    """
    columns = []
    for record in records:
        if not isinstance(record, dict):
            continue
        for key in record:
            if key not in columns:
                columns.append(key)
    return columns


def read_json(file_bytes: bytes):
    """JSON to rows, with the keys becoming the header row.

    Turning objects back into a header plus rows means JSON files go through
    exactly the same column-alias matching as CSV — 'Narration', 'narration'
    and 'description' all land in the same place, for free.
    """
    text = decode(file_bytes)

    try:
        payload = json.loads(text)
    except json.JSONDecodeError as error:
        raise UnreadableFile(f"File is not valid JSON: {error.msg}.")

    records = _records_from_json(payload)
    if not records:
        raise UnreadableFile("The JSON file contains no transactions.")

    # A list of lists is already rows; take it as-is.
    if all(isinstance(record, list) for record in records):
        return [[_flatten(cell) for cell in record] for record in records]

    columns = _columns_in_order(records)
    if not columns:
        raise UnreadableFile("The JSON records have no fields to read.")

    rows = [columns]
    for record in records:
        if not isinstance(record, dict):
            continue
        rows.append([_flatten(record.get(column, "")) for column in columns])
    return rows


def _flatten(value):
    """Cells must be scalars. A nested object becomes its JSON text.

    That keeps a weird file from crashing the import: the row simply fails
    to parse and gets counted as skipped, like any other bad row.
    """
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value


# --- excel ----------------------------------------------------------------

def read_excel_sheets(file_bytes: bytes):
    """Every worksheet of an .xlsx, as (name, rows) pairs.

    All of them, not just the first: banks and exporters routinely put one
    month per tab, and reading only sheet 1 silently loses the rest — the
    worst kind of bug, because the import still reports success.

    read_only keeps a large workbook from being built in memory all at once,
    and data_only asks for the cached result of a formula rather than the
    formula text, since '=SUM(D2:D9)' is not an amount.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UnreadableFile(
            "Excel support needs the openpyxl package. Install it, or export "
            "the statement as CSV."
        )

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as error:
        raise UnreadableFile(f"Could not open the Excel file: {error}")

    try:
        sheets = []
        for name in workbook.sheetnames:
            rows = [
                ["" if cell is None else cell for cell in row]
                for row in workbook[name].iter_rows(values_only=True)
            ]
            if rows:
                sheets.append((name, rows))
        return sheets
    finally:
        workbook.close()


def read_excel(file_bytes: bytes):
    """First worksheet only. Kept for callers that want a single table."""
    sheets = read_excel_sheets(file_bytes)
    return sheets[0][1] if sheets else []


# --- entry point ----------------------------------------------------------

READERS = {
    "delimited": read_delimited,
    "json": read_json,
    "excel": read_excel,
}


def read_sheets(file_bytes: bytes, filename=None):
    """Any supported file to a list of (sheet name, rows).

    Only Excel can hold more than one table, so the other formats return a
    single entry. Callers treat every format the same way and get multi-sheet
    support for nothing.
    """
    if not file_bytes:
        raise UnreadableFile("File is empty.")

    kind = detect_format(file_bytes, filename)

    if kind == "excel":
        sheets = read_excel_sheets(file_bytes)
    else:
        # One table, so there is no sheet to name. None rather than the
        # filename: the upload already records that, and repeating it would
        # show every CSV row as "statement.csv > statement.csv".
        sheets = [(None, READERS[kind](file_bytes))]

    sheets = [(name, rows) for name, rows in sheets if rows]
    if not sheets:
        raise UnreadableFile("File is empty.")
    return sheets


def read_rows(file_bytes: bytes, filename=None):
    """Any supported file to a list of rows, first table only."""
    return read_sheets(file_bytes, filename)[0][1]
