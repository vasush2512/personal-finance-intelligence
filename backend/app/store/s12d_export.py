"""Building export files: CSV and Excel (Phase 4).

Two decisions worth stating, because both are departures from how the rest of
the app handles the same data.

**Money is a number in the spreadsheet, not a string.** Everywhere else in this
project money leaves as a 2-decimal string, because JSON has no decimal type and
JavaScript would turn 409.50 into 409.49999999999994. A spreadsheet is the
opposite case: the entire reason someone exports to Excel is to sum a column,
and a column of text sums to zero. The value is quantised to paise before it is
written, so nothing is lost on the way out.

**Long digit runs are masked.** An export is the one place this data leaves the
application, and bank narrations carry card and account numbers. Nine or more
consecutive digits is a reference, a card or an account — never a merchant name
— so those are masked in the exported description. The merchant column beside it
is unaffected, so the row is still identifiable.
"""

import csv
import datetime as dt
import io
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Nine digits is past any plausible amount or year and well into card, account
# and UTR territory. Four would also catch a 2026 in a narration.
_LONG_DIGITS = re.compile(r"\d{9,}")

TRANSACTION_HEADERS = [
    "Date",
    "Description",
    "Merchant",
    "Payment method",
    "Amount",
    "Direction",
    "Category",
    "Labelled by",
    "Model confidence",
    "Source file",
    "Worksheet",
]


def mask_identifiers(text: str) -> str:
    """Replace long digit runs with their last four digits.

    '...412345678901/SWIGGY...' -> '...XXXX8901/SWIGGY...'. Keeping the tail
    means a row can still be matched against a bank statement by eye, which is
    the main reason anyone exports one.
    """
    if not text:
        return ""
    return _LONG_DIGITS.sub(lambda match: f"XXXX{match.group()[-4:]}", str(text))


def transaction_rows(transactions, upload_names=None):
    """Model rows -> lists of cell values, in TRANSACTION_HEADERS order."""
    upload_names = upload_names or {}

    from app.pipeline.s05_normalize import extract_payment_method

    return [
        [
            row.date,
            mask_identifiers(row.description),
            row.merchant,
            extract_payment_method(row.description) or "",
            Decimal(row.amount).quantize(Decimal("0.01")),
            row.direction,
            row.category,
            row.category_source,
            round(row.confidence, 3) if row.confidence is not None else "",
            upload_names.get(row.upload_id, ""),
            row.sheet_name or "",
        ]
        for row in transactions
    ]


def to_csv(headers, rows) -> bytes:
    """UTF-8 with a BOM, so Excel opens ₹ and merchant names correctly.

    Without the BOM, Excel on Windows reads a UTF-8 CSV as the system codepage
    and every non-ASCII character in a merchant name becomes mojibake.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(headers)

    for row in rows:
        writer.writerow(
            [
                # A date written by str() sorts and parses correctly everywhere;
                # a locale-formatted one does neither.
                value.isoformat() if isinstance(value, dt.date) else
                f"{value:.2f}" if isinstance(value, Decimal) else value
                for value in row
            ]
        )

    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_excel(sheets) -> bytes:
    """`sheets` is [(title, headers, rows), ...] -> one workbook.

    Amounts are written as numbers so the column can be summed. Dates are
    written as dates for the same reason.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title, headers, rows in sheets:
        # Excel refuses sheet names over 31 characters or containing []:*?/\
        sheet = workbook.create_sheet(re.sub(r"[\[\]:*?/\\]", "-", title)[:31])
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append(row)

        _format_columns(sheet, headers, rows)
        # Freeze the header, so scrolling 100,000 rows stays readable.
        sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _format_columns(sheet, headers, rows):
    """Number formats and widths, from what the first rows actually contain."""
    sample = rows[:200]

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        values = [row[index - 1] for row in sample if len(row) >= index]

        if any(isinstance(value, Decimal) for value in values):
            for cell in sheet[letter][1:]:
                cell.number_format = "#,##0.00"
        elif any(isinstance(value, dt.date) for value in values):
            for cell in sheet[letter][1:]:
                cell.number_format = "yyyy-mm-dd"

        longest = max(
            [len(str(header))] + [len(str(value)) for value in values[:50]]
        )
        sheet.column_dimensions[letter].width = min(max(longest + 2, 10), 60)


def summary_sheets(summary, trends):
    """The analytical export: categories, months and merchants."""
    return [
        (
            "Categories",
            ["Category", "Transactions", "Total"],
            [
                [row["category"], row["count"], Decimal(str(row["total"]))]
                for row in summary["by_category"]
            ],
        ),
        (
            "Monthly",
            ["Month", "Spent", "Income"],
            [
                [
                    point["month"],
                    Decimal(str(point["spent"])),
                    Decimal(str(point["income"])),
                ]
                for point in trends
            ],
        ),
        (
            "Merchants",
            ["Merchant", "Transactions", "Total"],
            [
                [row["merchant"], row["count"], Decimal(str(row["total"]))]
                for row in summary["top_merchants"]
            ],
        ),
    ]


def filename(kind: str, extension: str, today=None) -> str:
    """'expense-tracker-transactions-2026-08-15.csv'."""
    today = today or dt.date.today()
    return f"expense-tracker-{kind}-{today.isoformat()}.{extension}"
