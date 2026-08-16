"""CSV and Excel export (Phase 4).

The two things most worth pinning: an amount must arrive in a spreadsheet as
something you can sum, and an account number must not arrive at all.
"""

import csv
import datetime as dt
import io
from decimal import Decimal

from openpyxl import load_workbook

from app.store.s12d_export import (
    TRANSACTION_HEADERS,
    filename,
    mask_identifiers,
    summary_sheets,
    to_csv,
    to_excel,
    transaction_rows,
)


class FakeRow:
    """A transaction as the exporter reads it, without needing a database."""

    def __init__(self, description, amount="1450.00", category="food",
                 source="rule", confidence=None, upload_id=1, sheet=None):
        self.date = dt.date(2026, 6, 12)
        self.description = description
        self.merchant = "Swiggy"
        self.amount = Decimal(amount)
        self.direction = "debit"
        self.category = category
        self.category_source = source
        self.confidence = confidence
        self.upload_id = upload_id
        self.sheet_name = sheet


def read_csv(body):
    text = body.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


# --- masking ---------------------------------------------------------------


def test_a_long_reference_number_is_masked_to_its_last_four():
    assert mask_identifiers("UPI/DR/412345678901/SWIGGY") == "UPI/DR/XXXX8901/SWIGGY"


def test_a_short_number_is_left_alone():
    """A year or an amount in a narration is not an account number."""
    assert mask_identifiers("RENT MARCH 2026") == "RENT MARCH 2026"
    assert mask_identifiers("POS 4512 CAFE") == "POS 4512 CAFE"


def test_the_merchant_column_survives_masking():
    rows = transaction_rows([FakeRow("UPI/DR/412345678901/SWIGGY/HDFC")])
    description, merchant = rows[0][1], rows[0][2]
    assert "412345678901" not in description
    assert merchant == "Swiggy"


def test_masking_happens_on_export_not_in_the_database():
    """The stored row is untouched; only the exported copy is masked."""
    row = FakeRow("UPI/DR/412345678901/SWIGGY")
    transaction_rows([row])
    assert row.description == "UPI/DR/412345678901/SWIGGY"


# --- CSV -------------------------------------------------------------------


def test_the_csv_starts_with_a_bom_so_excel_reads_utf8():
    body = to_csv(["A"], [["x"]])
    assert body.startswith(b"\xef\xbb\xbf")


def test_the_csv_header_matches_the_columns():
    body = to_csv(TRANSACTION_HEADERS, transaction_rows([FakeRow("SWIGGY")]))
    assert read_csv(body)[0] == TRANSACTION_HEADERS


def test_money_keeps_both_decimal_places_in_csv():
    body = to_csv(TRANSACTION_HEADERS, transaction_rows([FakeRow("SWIGGY", "1450.00")]))
    amount = read_csv(body)[1][TRANSACTION_HEADERS.index("Amount")]
    assert amount == "1450.00"


def test_dates_are_written_iso_so_they_sort_and_parse_anywhere():
    body = to_csv(TRANSACTION_HEADERS, transaction_rows([FakeRow("SWIGGY")]))
    assert read_csv(body)[1][0] == "2026-06-12"


def test_a_missing_confidence_is_blank_not_zero():
    """Zero would read as 'the model was certain it was wrong'."""
    body = to_csv(TRANSACTION_HEADERS, transaction_rows([FakeRow("SWIGGY")]))
    assert read_csv(body)[1][TRANSACTION_HEADERS.index("Model confidence")] == ""


# --- Excel -----------------------------------------------------------------


def sheet_from(body, index=0):
    workbook = load_workbook(io.BytesIO(body))
    return workbook[workbook.sheetnames[index]]


def test_an_amount_arrives_in_excel_as_a_number_you_can_sum():
    """A column of text sums to zero, which defeats the point of exporting."""
    body = to_excel([("Transactions", TRANSACTION_HEADERS,
                      transaction_rows([FakeRow("SWIGGY", "1450.00")]))])
    cell = sheet_from(body).cell(row=2, column=TRANSACTION_HEADERS.index("Amount") + 1)
    assert isinstance(cell.value, (int, float, Decimal))
    assert float(cell.value) == 1450.00


def test_paise_survive_the_trip_into_a_spreadsheet():
    """Money is the one thing an export must not round. A cell that reads
    732.55 in the app and 732.60 in Excel makes the whole file useless."""
    amounts = [Decimal("732.55"), Decimal("0.01"), Decimal("1450.00"),
               Decimal("123456.78")]
    body = to_excel([("T", ["Amount"], [[amount] for amount in amounts])])
    sheet = sheet_from(body)

    written = [sheet.cell(row=index, column=1).value
               for index in range(2, sheet.max_row + 1)]
    assert [Decimal(str(value)) for value in written] == amounts
    # And every one still displays to two places, including the whole rupee.
    assert all(
        sheet.cell(row=index, column=1).number_format == "#,##0.00"
        for index in range(2, sheet.max_row + 1)
    )


def test_the_header_row_is_frozen_and_bold():
    body = to_excel([("Transactions", TRANSACTION_HEADERS,
                      transaction_rows([FakeRow("SWIGGY")]))])
    sheet = sheet_from(body)
    assert sheet.freeze_panes == "A2"
    assert sheet.cell(row=1, column=1).font.bold


def test_the_summary_export_has_a_sheet_per_table():
    summary = {
        "by_category": [{"category": "food", "total": "20000.00", "count": 40}],
        "top_merchants": [{"merchant": "Swiggy", "total": "9000.00", "count": 22}],
    }
    trends = [{"month": "2026-06", "spent": "50000.00", "income": "80000.00"}]

    body = to_excel(summary_sheets(summary, trends))
    names = load_workbook(io.BytesIO(body)).sheetnames
    assert names == ["Categories", "Monthly", "Merchants"]


def test_a_sheet_name_excel_would_reject_is_cleaned():
    body = to_excel([("June/July [2026]: totals", ["A"], [["x"]])])
    name = load_workbook(io.BytesIO(body)).sheetnames[0]
    assert not set(name) & set("[]:*?/\\")
    assert len(name) <= 31


def test_an_empty_export_still_produces_a_valid_file():
    """Nobody should get a corrupt download because a filter matched nothing."""
    body = to_excel([("Transactions", TRANSACTION_HEADERS, [])])
    sheet = sheet_from(body)
    assert [cell.value for cell in sheet[1]] == TRANSACTION_HEADERS
    assert sheet.max_row == 1

    rows = read_csv(to_csv(TRANSACTION_HEADERS, []))
    assert rows == [TRANSACTION_HEADERS]


# --- naming ----------------------------------------------------------------


def test_the_filename_carries_the_date_it_was_taken():
    assert filename("transactions", "csv", today=dt.date(2026, 8, 15)) == (
        "expense-tracker-transactions-2026-08-15.csv"
    )
